import openpyxl

def load_on_excel(diccionary:dict,company:int,product:str,flag:bool):
    """
    Esta funcion es la primera de una serie de funciones que se encargan de realizar la carga de la request en el excel
    - diccionary: Lista de precios que queremos escribir
    - company: identificador unico de la empresa de la cual queremos obtener informacion [ FULLHARD: 1 | VENEX: 2 | GEZATEK: 3 ]
    - product: Identificador del producto que queremos obtener la lista de precios(Procesadores AMD) - (Procesadores Intel) -
        (Motherboards AMD) - (Motherboards Intel) - (Memorias RAM) - (SSD) - (HDD)
    - flag: Booleano que nos permite ejecutar el programa de manera diferente durante la primer escritura
    """
    
    # Creamos el arhcivo de excel
    if flag:
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook("Lista_de_precios.xlsx")
    
    if company == 1:
        name_column = 1
        price_column = 2
    elif company == 2:
        name_column = 4
        price_column = 5
    elif company == 3:
        name_column = 7
        price_column = 8
    else:
        raise ValueError("--El número ingresado no corresponde con ninguna empresa dentro de nuestra base de datos--")
    
    # Creamos una hoja
    hoja =  create_sheets_price_list(workbook,product,flag)
    
    write_price_list(diccionary,name_column,price_column,hoja)
    
    workbook.save('Lista_de_precios.xlsx')
    
    return flag


def write_price_list(diccionary:dict,name_column:int,price_column:int,hoja):
    """
    Esta funcion se encarga de escribir y guardar en un archivo de excel la lista de precios que deseamos
    - diccionary: Lista de precios que queremos escribir
    - name_column : Numero que indica el numero de la columna en la cual queremos escribir 
    - price_colum : Numero que indica el numero de la cual columna en la cual queremos escribir 
    - hoja : Hoja en la cual vamos a hacer la escritura
    """

    try: 
        fila = 2
        for valor in diccionary.keys():
            hoja.cell(row=fila, column=name_column, value=valor)
            fila += 1

        fila = 2
        for valor in diccionary.values():
            hoja.cell(row=fila, column=price_column, value=valor)
            fila += 1
            
    except:
        print("--Sucedio un error durante la carga la lista!--")
        

def create_sheets_price_list(workbook,product:str,flag:bool):
    """
    Esta funcion se encarga de aplicarle ciertas caracteristicas a cada una de las hojas y de direccionar la escritura a la hoja correspondiente 
    - workbook: Contenedor 
    - product: Identificador del producto que queremos obtener la lista de precios(Procesadores AMD) - (Procesadores Intel) -
        (Motherboards AMD) - (Motherboards Intel) - (Memorias RAM) - (SSD) - (HDD)
    - flag: Booleano que nos permite ejecutar el programa de manera diferente durante la primer escritura
    """
    
    if flag:
        sheet_list = ["Procesadores AMD","Procesadores Intel","Motherboards AMD","Motherboards Intel",
                        "Memorias RAM","SSD","HDD"]
        for i in sheet_list:
            # Aplicamos ciertas propiedades a cada una de las hojas 
            hoja = workbook.create_sheet(i)
            hoja['A1'] = 'COMPRAGAMER'
            hoja['D1'] = 'VENEX'
            hoja['G1'] = 'GEZATEK'
            hoja.column_dimensions['A'].width = 80
            hoja.column_dimensions['D'].width = 80
            hoja.column_dimensions['G'].width = 80
        workbook.remove(workbook.get_sheet_by_name("Sheet"))
        flag = not flag # Cambiamos la bandera para que no se vuelva a ingresar a aesta parte del codgio 
    
    try:
        match product:
            case "Procesadores AMD":
                hoja = "Procesadores AMD"
            case "Procesadores Intel":
                hoja = "Procesadores Intel"
            case "Motherboards AMD":
                hoja = "Motherboards AMD"
            case "Motherboards Intel":
                hoja = "Motherboards Intel"
            case "Memorias RAM":
                hoja = "Memorias RAM"
            case "SSD":
                hoja = "SSD"
            case "HDD":
                hoja = "HDD"
            case _:
                raise ValueError
            
        return workbook.get_sheet_by_name(hoja)
    
    except ValueError:
        print("--El producto ingresado es INCORRECTO--")
        
